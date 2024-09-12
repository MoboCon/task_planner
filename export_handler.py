import csv
import json
from fpdf import FPDF
import openpyxl

def export_to_csv(tasks):
    with open('tasks_export.csv', mode='w', newline='', encoding='utf-8') as file:
        writer = csv.writer(file)
        writer.writerow(["Titlu", "Categorie", "Subcategorie", "Data limită", "Prioritate", "Stare"])
        for task in tasks:
            writer.writerow([task[1], task[3], task[4], task[5], task[6], task[7]])

def export_to_pdf(tasks):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)

    pdf.cell(200, 10, txt="Task Planner - Export PDF", ln=True, align="C")
    pdf.ln(10)

    for task in tasks:
        pdf.cell(200, 10, txt=f"Titlu: {task[1]}, Categorie: {task[3]}, Subcategorie: {task[4]}", ln=True)

    pdf.output("tasks_export.pdf")

def export_to_excel(tasks):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sarcini"
    
    sheet.append(["Titlu", "Categorie", "Subcategorie", "Data limită", "Prioritate", "Stare"])
    for task in tasks:
        sheet.append([task[1], task[3], task[4], task[5], task[6], task[7]])

    workbook.save("tasks_export.xlsx")

def export_to_json(tasks):
    tasks_data = [{"Titlu": task[1], "Categorie": task[3], "Subcategorie": task[4], "Data limită": task[5], "Prioritate": task[6], "Stare": task[7]} for task in tasks]
    
    with open('tasks_export.json', 'w') as f:
        json.dump(tasks_data, f, indent=4)
